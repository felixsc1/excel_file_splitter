import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl


def set_integer_format(sheet, dataframe):
    """
    Apply integer format to columns in the given dataframe that have integer values.
    """
    for col_idx, (col, col_data) in enumerate(dataframe.items(), start=1):
        try:
            # Check if all non-NaN values are integers
            if (
                col_data.dropna()
                .apply(lambda x: isinstance(x, (int, float)) and float(x).is_integer())
                .all()
            ):
                for col_cells in sheet.iter_cols(min_col=col_idx, max_col=col_idx):
                    for cell in col_cells:
                        cell.number_format = "0"
        except ValueError:
            continue


def copy_cell(old_cell, new_cell):
    """Copy the value, style, and hyperlink of old_cell to new_cell."""
    new_cell.value = old_cell.value
    if old_cell.has_style:
        new_cell.font = old_cell.font.copy()
        new_cell.border = old_cell.border.copy()
        new_cell.fill = old_cell.fill.copy()
        new_cell.number_format = old_cell.number_format
        new_cell.protection = old_cell.protection.copy()
        new_cell.alignment = old_cell.alignment.copy()
    if old_cell.hyperlink:
        new_cell.hyperlink = old_cell.hyperlink


def get_or_create_sheet(wb, sheet_name, position=None):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    else:
        sheet = wb.create_sheet(sheet_name)
        if position is not None:
            wb._sheets.sort(
                key=lambda ws: position
                if ws.title == sheet_name
                else wb.sheetnames.index(ws.title)
            )
        return sheet


def process_sheet(
    old_workbook,
    workbook,
    sheet_name,
    dataframe,
    start_row=None,
    skip_header=False,
    position=None,
):
    old_ws = old_workbook[sheet_name]
    ws = get_or_create_sheet(workbook, sheet_name, position=position)

    # Copy the header if not skipped
    if not skip_header:
        for j, col in enumerate(old_ws.iter_cols(min_row=1, max_row=1)):
            new_cell = ws.cell(row=1, column=j + 1)
            copy_cell(col[0], new_cell)

    row_start = 2 if not skip_header else 1

    # Copy the data rows
    rows_to_process = dataframe_to_rows(dataframe, index=False, header=False)
    for i, row_df in enumerate(rows_to_process):
        row_offset = start_row + i if start_row else i
        for j, value in enumerate(row_df):
            cell = ws.cell(row=i + row_start, column=j + 1)
            old_cell = old_ws.cell(row=row_offset + 1, column=j + 1)
            copy_cell(old_cell, cell)


def copy_whole_sheet(source_workbook, target_workbook, sheet_name, position=None):
    source_sheet = source_workbook[sheet_name]

    if sheet_name in target_workbook.sheetnames:
        target_sheet = target_workbook[sheet_name]
    else:
        if position is None:
            target_sheet = target_workbook.create_sheet(sheet_name)
        else:
            target_sheet = target_workbook.create_sheet(sheet_name, position)

    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(
                row=cell.row, column=cell.col_idx, value=cell.value
            )
            copy_cell(cell, new_cell)


def split_excel(input_file, rows_per_file, suffix_template):
    original_workbook = load_workbook(input_file)
    # Read all sheets from the excel file
    with pd.ExcelFile(input_file) as xls:
        sheet_names = xls.sheet_names  # Capture the order of sheets
        all_sheets = {sheet_name: xls.parse(sheet_name) for sheet_name in sheet_names}

    # Identify the sheet with the maximum rows
    max_rows_sheet_name = max(all_sheets, key=lambda sheet: len(all_sheets[sheet]))
    max_rows_sheet = all_sheets[max_rows_sheet_name]

    # Calculate number of smaller files required for the max_rows_sheet
    no_of_files = len(max_rows_sheet) // rows_per_file + (
        1 if len(max_rows_sheet) % rows_per_file else 0
    )

    # Determine the output directory and original filename from the input file's path
    output_dir = os.path.dirname(input_file)
    original_filename = os.path.splitext(os.path.basename(input_file))[
        0
    ]  # get filename without extension

    for i in range(no_of_files):
        # Extract subset of rows for the main sheet
        start_row = i * rows_per_file
        end_row = (i + 1) * rows_per_file
        subset_main_sheet = max_rows_sheet[start_row:end_row]

        # Create a new excel writer for this subset using the xlsxwriter engine
        formatted_number = str(i + 1).zfill(3)  # this will produce '001', '002', etc.
        output_suffix = suffix_template.replace("{number}", formatted_number)
        output_file = os.path.join(
            output_dir, f"{original_filename}_{output_suffix}.xlsx"
        )

        with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:
            # Write data to the Excel file
            subset_main_sheet.to_excel(
                writer, sheet_name=max_rows_sheet_name, index=False
            )
            for sheet_name in sheet_names:
                if sheet_name != max_rows_sheet_name:
                    all_sheets[sheet_name].to_excel(
                        writer, sheet_name=sheet_name, index=False
                    )

        # Now, use openpyxl to post-process the Excel file
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        for sheet_name in sheet_names:
            wb.create_sheet(sheet_name)

        # Processing the max_rows_sheet_name sheet for hyperlink and style copying
        old_ws = original_workbook[max_rows_sheet_name]
        ws = wb[max_rows_sheet_name]
        for i, row_df in enumerate(
            dataframe_to_rows(subset_main_sheet, index=False, header=True)
        ):
            for j, value in enumerate(row_df):
                cell = ws.cell(row=i + 1, column=j + 1)
                old_cell = old_ws.cell(row=start_row + i + 1, column=j + 1)
                copy_cell(old_cell, cell)

        for sheet_name in sheet_names:
            position = original_workbook.sheetnames.index(sheet_name)
            if sheet_name == max_rows_sheet_name:
                process_sheet(
                    original_workbook,
                    wb,
                    sheet_name,
                    subset_main_sheet,
                    start_row,
                    position=position,
                )
            else:
                copy_whole_sheet(original_workbook, wb, sheet_name, position)

        # Post-process formatting
        # wb = writer.book  # this is the new book we're writing to
        # sheet = wb[max_rows_sheet_name]
        sheet = wb.active
        set_integer_format(sheet, subset_main_sheet)
        for sheet_name in sheet_names:
            if sheet_name != max_rows_sheet_name:
                set_integer_format(wb[sheet_name], all_sheets[sheet_name])

        wb.save(output_file)
        print(f"Wrote file {output_file}")


def main_gui():
    root = tk.Tk()
    root.title("Excel Splitter")

    def select_input_file():
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        input_file_var.set(file_path)

    def split_and_save():
        input_file = input_file_var.get()
        rows = int(rows_per_file_var.get())
        suff_template = suffix_var.get()

        split_excel(input_file, rows, suff_template)
        messagebox.showinfo("Success", "Excel file split successfully!")

    # GUI Elements
    input_file_var = tk.StringVar()
    rows_per_file_var = tk.StringVar(value="1000")  # Default value set to 1000
    suffix_var = tk.StringVar(value="{number}")  # Default value

    tk.Label(root, text="Select Excel File:").pack(pady=20)
    tk.Entry(root, textvariable=input_file_var, width=50).pack(pady=5)
    tk.Button(root, text="Browse", command=select_input_file).pack(pady=10)

    tk.Label(root, text="Rows per file:").pack(pady=20)
    tk.Entry(root, textvariable=rows_per_file_var, width=50).pack(pady=5)

    tk.Label(root, text="Suffix for split files (use {number} for numbering):").pack(
        pady=20
    )
    tk.Entry(root, textvariable=suffix_var, width=50).pack(pady=5)

    tk.Button(root, text="Split and Save", command=split_and_save).pack(pady=20)

    root.mainloop()


if __name__ == "__main__":
    main_gui()
